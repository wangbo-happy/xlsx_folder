#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel文件分析模块
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple


def analyze_excel_file(file_path: str) -> Dict:
    """
    分析Excel文件内容
    
    Args:
        file_path: Excel文件路径
        
    Returns:
        包含分析结果的字典
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件 {file_path} 不存在")
        
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"无法打开Excel文件: {e}")
    
    result = {
        'file_name': os.path.basename(file_path),
        'sheets': [],
        'total_sheets': len(wb.sheetnames),
        'file_size': os.path.getsize(file_path)
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 获取工作表维度
        dimensions = ws.calculate_dimension()
        
        # 统计非空单元格数量
        non_empty_cells = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
        
        sheet_data = {
            'sheet_name': sheet_name,
            'dimensions': dimensions,
            'non_empty_cells': non_empty_cells,
            'max_row': ws.max_row,
            'max_column': ws.max_column
        }
        
        result['sheets'].append(sheet_data)
    
    wb.close()
    return result


if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print("用法: python analyze_excel.py <excel文件路径>")
        sys.exit(1)
        
    file_path = sys.argv[1]
    
    try:
        analysis = analyze_excel_file(file_path)
        print(f"文件分析结果:\n{'-'*30}")
        print(f"文件名: {analysis['file_name']}")
        print(f"工作表数量: {analysis['total_sheets']}")
        print(f"文件大小: {analysis['file_size']} 字节")
        
        for sheet in analysis['sheets']:
            print(f"\n工作表: {sheet['sheet_name']}")
            print(f"数据范围: {sheet['dimensions']}")
            print(f"行数: {sheet['max_row']}")
            print(f"列数: {sheet['max_column']}")
            print(f"非空单元格: {sheet['non_empty_cells']}")
            
    except Exception as e:
        print(f"错误: {e}")